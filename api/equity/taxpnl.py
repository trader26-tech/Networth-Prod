"""Parse a Zerodha **Tax P&L** export and pull out the numbers that drive
long-term-capital-gains (LTCG) harvesting.

Every financial year an individual gets ₹1,25,000 of LTCG on listed equity that
is TAX-FREE (Sec 112A). To harvest it you want to know, per account:
  • how much LONG-TERM realised gain you have already booked this FY, and
  • how much of the ₹1.25 L exemption is therefore still unused.

The right file is the **Tax P&L** (Console → Reports → Tax P&L), NOT the plain
"P&L" statement — only the Tax P&L splits realised profit into Long Term / Short
Term / Intraday and carries a clear FY period. We read the summary block on the
"Equity and Non Equity" sheet:

    Realized Profit Breakdown
      Intraday/Speculative profit   0.0
      Short Term profit             51787.13
      Long Term profit              0.0          ← the LTCG number we harvest against
      Non Equity profit             5134.86

and the period from the sheet title ("… from 2025-04-01 to 2026-03-31"), which we
check against the FY the user entered so a wrong file can't be filed silently.
"""
from __future__ import annotations

import io
import re
from typing import Any, Optional

# The Sec 112A tax-free LTCG allowance per person per FY (FY2024-25 onward).
LTCG_FREE_ALLOWANCE = 125000.0
# Statutory rates on LISTED, STT-paid equity for FY2024-25 onward (Budget 2024):
#   • STCG (Sec 111A): flat 20%
#   • LTCG (Sec 112A): 12.5% on the gain ABOVE the ₹1.25 L free allowance
STCG_RATE = 0.20
LTCG_RATE = 0.125

_SUMMARY_SHEET_HINTS = ("equity and non equity", "equity")
_PERIOD_RE = re.compile(r"from\s*(\d{4}-\d{2}-\d{2})\s*to\s*(\d{4}-\d{2}-\d{2})", re.I)


class TaxPnlError(ValueError):
    """A human-readable reason the file couldn't be used (wrong file, wrong FY…)."""


def _norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


# Zerodha's Tax P&L labels the equity summary two ways depending on the export:
#   old: "Long Term profit" / "Short Term profit" / "Intraday/Speculative profit"
#   new: "Equity Long Term profit" / "Equity Short Term profit" / "Equity Intraday…"
# Match either, but NOT the Mutual-Funds sheet's look-alikes, which put the asset
# class as a SUFFIX ("Long Term profit Equity", "… Debt") — those are a different
# number and would corrupt the LTCG figure.
def _classify_summary_label(t: str) -> Optional[str]:
    """Return 'long term' | 'short term' | 'intraday' | 'non equity' for an equity
    realised-profit summary label, else None."""
    if not t:
        return None
    # drop a leading "equity " qualifier (the new format); keep the rest intact
    body = t[len("equity "):].strip() if t.startswith("equity ") else t
    # MF suffix forms end in " equity" / " debt" → not the equity-cash summary
    if body.endswith(" equity") or body.endswith(" debt"):
        return None
    if body == "long term profit":
        return "long term"
    if body == "short term profit":
        return "short term"
    if body in ("intraday/speculative profit", "intraday profit", "speculative profit"):
        return "intraday"
    if body == "non equity profit":
        return "non equity"
    return None


def _num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def fy_bounds(fy: str) -> tuple[str, str, str]:
    """'2025-26' | '2025-2026' | 'FY2025-26' | '2025' → ('2025-26', '2025-04-01',
    '2026-03-31'). Raises TaxPnlError on anything unparseable."""
    m = re.search(r"(20\d{2})\s*[-/ ]\s*(\d{2,4})?", str(fy or ""))
    if not m:
        m2 = re.search(r"(20\d{2})", str(fy or ""))
        if not m2:
            raise TaxPnlError(f"Couldn't read a financial year from “{fy}”. Use e.g. 2025-26.")
        start = int(m2.group(1))
    else:
        start = int(m.group(1))
    label = f"{start}-{str(start + 1)[-2:]}"
    return label, f"{start}-04-01", f"{start + 1}-03-31"


def fy_from_period(parsed: dict) -> str:
    """The FY label a parsed file belongs to, taken from its OWN period. Indian FY
    runs Apr→Mar, so the start year of period_from is the FY start. Falls back to
    period_to's year − 1 if the start is missing. Raises if neither is present."""
    pf = str(parsed.get("period_from") or "")
    pt = str(parsed.get("period_to") or "")
    m = re.match(r"(\d{4})-(\d{2})-\d{2}", pf)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        start = y if mo >= 4 else y - 1        # Jan–Mar belongs to the previous FY
        return f"{start}-{str(start + 1)[-2:]}"
    m2 = re.match(r"(\d{4})-(\d{2})-\d{2}", pt)
    if m2:
        y, mo = int(m2.group(1)), int(m2.group(2))
        end = y if mo <= 3 else y + 1          # period_to is the FY end
        return f"{end - 1}-{str(end)[-2:]}"
    raise TaxPnlError(
        "The file has no date range, so its financial year can't be determined. "
        "Re-download the Tax P&L from Zerodha Console for a specific FY.")


def _read_workbook(data: bytes):
    try:
        import openpyxl
    except Exception as e:                       # pragma: no cover
        raise TaxPnlError(f"Server can't read .xlsx files ({e}).")
    try:
        # NOT read_only: Zerodha's sheets carry unreliable dimension metadata that
        # makes read_only iteration return truncated/empty rows.
        return openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise TaxPnlError(
            "That doesn't look like an .xlsx Tax P&L. Download it from Zerodha Console "
            f"→ Reports → Tax P&L (not the plain P&L), as Excel. ({e})")


def _pick_summary_sheet(wb):
    names = {_norm(n): n for n in wb.sheetnames}
    for hint in _SUMMARY_SHEET_HINTS:
        for norm, real in names.items():
            if norm == hint:
                return wb[real]
    for norm, real in names.items():             # loose contains-match fallback
        if "equity" in norm and "non equity" in norm:
            return wb[real]
    # last resort: the first sheet that actually has a "Long Term profit" label
    for real in wb.sheetnames:
        ws = wb[real]
        mc = max(ws.max_column or 1, 16)
        for r in range(1, 41):
            if any(_classify_summary_label(_norm(ws.cell(r, c).value)) == "long term"
                   for c in range(1, mc + 1)):
                return ws
    return None


def parse(data: bytes) -> dict:
    """Extract the harvesting-relevant totals from a Tax P&L .xlsx.

    Returns {period_from, period_to, fy_label, long_term, short_term, intraday,
    non_equity, realized_total, source_sheet, client_id?}. Raises TaxPnlError with
    a clear message when the file is the wrong kind."""
    wb = _read_workbook(data)
    ws = _pick_summary_sheet(wb)
    if ws is None:
        raise TaxPnlError(
            "This looks like a plain P&L, not a Tax P&L — it has no Long/Short-term "
            "split. In Zerodha Console open Reports → Tax P&L and download that.")

    period_from = period_to = None
    client_id = None
    vals: dict[str, float] = {}
    max_c = max(ws.max_column or 1, 16)
    # scan the summary block (top ~60 rows) for the labelled numbers + period title.
    # Read by explicit coordinates — Zerodha's sheet dimensions are unreliable, so
    # iter_rows can under-report the used range.
    for r in range(1, 61):
        cells = [ws.cell(r, c).value for c in range(1, max_c + 1)]
        for i, c in enumerate(cells):
            t = _norm(c)
            if not t:
                continue
            if period_from is None:
                m = _PERIOD_RE.search(str(c))
                if m:
                    period_from, period_to = m.group(1), m.group(2)
            if t == "client id" and i + 1 < len(cells):
                client_id = str(cells[i + 1] or "").strip() or None
            key = _classify_summary_label(t)
            if key and key not in vals:      # first (equity-summary) occurrence wins
                # value is the next non-empty cell on the row
                for c2 in cells[i + 1:]:
                    n = _num(c2)
                    if n is not None:
                        vals[key] = n
                        break

    if "long term" not in vals and "short term" not in vals:
        raise TaxPnlError(
            "Couldn't find the Long/Short-term profit summary. Make sure this is the "
            "Tax P&L export (Console → Reports → Tax P&L), not the plain P&L.")

    long_term = vals.get("long term", 0.0)
    short_term = vals.get("short term", 0.0)
    intraday = vals.get("intraday", 0.0)
    non_equity = vals.get("non equity", 0.0)

    return {
        "period_from": period_from,
        "period_to": period_to,
        "client_id": client_id,
        "long_term": round(long_term, 2),
        "short_term": round(short_term, 2),
        "intraday": round(intraday, 2),
        "non_equity": round(non_equity, 2),
        "realized_total": round(long_term + short_term + intraday + non_equity, 2),
        "source_sheet": ws.title,
    }


def harvest_view(parsed: dict, fy: str) -> dict:
    """Turn a parsed file + the FY the user entered into the harvesting numbers,
    including a period-vs-FY sanity check. Never raises for a period mismatch —
    it flags it (fy_mismatch) so the UI can warn but still show the numbers."""
    label, fy_from, fy_to = fy_bounds(fy)
    booked_lt = float(parsed.get("long_term") or 0.0)
    booked_st = float(parsed.get("short_term") or 0.0)
    # only a positive long-term gain uses up the tax-free room; a net LT loss leaves
    # the whole allowance intact (and is itself carry-forwardable, but that's separate)
    used = max(0.0, booked_lt)
    remaining = max(0.0, LTCG_FREE_ALLOWANCE - used)

    # ── tax already accrued this FY on the realised gains ──
    # STCG: flat 20% on a positive short-term gain (a loss owes nothing).
    st_taxable = max(0.0, booked_st)
    st_tax = round(st_taxable * STCG_RATE, 2)
    # LTCG: 12.5% only on the long-term gain ABOVE the ₹1.25 L allowance.
    lt_taxable = max(0.0, used - LTCG_FREE_ALLOWANCE)
    lt_tax = round(lt_taxable * LTCG_RATE, 2)
    total_tax = round(st_tax + lt_tax, 2)

    pf, pt = parsed.get("period_from"), parsed.get("period_to")
    mismatch = bool(pf and pt) and not (pf == fy_from and pt == fy_to)

    return {
        "fy_label": label,
        "fy_from": fy_from,
        "fy_to": fy_to,
        "allowance": LTCG_FREE_ALLOWANCE,
        "lt_booked": round(booked_lt, 2),
        "st_booked": round(booked_st, 2),
        "lt_used": round(used, 2),
        "lt_remaining": round(remaining, 2),
        "fully_used": remaining <= 0.005,
        # tax owed on what's already booked (listed equity, FY2024-25+ rates)
        "st_rate": STCG_RATE,
        "lt_rate": LTCG_RATE,
        "st_taxable": round(st_taxable, 2),
        "st_tax": st_tax,
        "lt_taxable": round(lt_taxable, 2),
        "lt_tax": lt_tax,
        "total_tax": total_tax,
        "period_from": pf,
        "period_to": pt,
        "fy_mismatch": mismatch,
        "mismatch_note": (
            f"This file covers {pf} → {pt}, but you entered FY {label} "
            f"({fy_from} → {fy_to}). Check you picked the right FY when exporting."
            if mismatch else None),
    }
