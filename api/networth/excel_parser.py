"""Excel parsing helpers for the Net-Worth importer.

Built on openpyxl (pandas is unavailable in this venv). Handles large,
messy multi-sheet workbooks (e.g. Numbers exports with 250+ sheets) by:
  - listing every sheet with a small preview + a header-row guess
  - returning a JSON-safe grid for a single sheet (capped)
All datetimes are serialised to ISO date strings so the payload is JSON-safe.
"""
from __future__ import annotations

import datetime as _dt
from typing import Any

import openpyxl


# --------------------------------------------------------------------------
# Cell serialisation
# --------------------------------------------------------------------------
def cell_to_json(v: Any) -> Any:
    """Convert an openpyxl cell value into a JSON-serialisable primitive."""
    if v is None:
        return None
    if isinstance(v, (_dt.datetime, _dt.date)):
        try:
            return v.date().isoformat() if isinstance(v, _dt.datetime) else v.isoformat()
        except Exception:
            return str(v)
    if isinstance(v, _dt.time):
        return v.isoformat()
    if isinstance(v, float):
        # collapse integral floats (217768763.0 -> 217768763)
        if v.is_integer():
            return int(v)
        return round(v, 6)
    if isinstance(v, (int, str, bool)):
        return v
    return str(v)


def _is_num(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _trim_trailing(row: list) -> list:
    r = list(row)
    while r and (r[-1] is None or (isinstance(r[-1], str) and r[-1].strip() == "")):
        r.pop()
    return r


def _row_has_content(row: list) -> bool:
    return any(c is not None and (not isinstance(c, str) or c.strip() != "") for c in row)


# --------------------------------------------------------------------------
# Sheet listing (fast, read-only)
# --------------------------------------------------------------------------
def list_sheets(path: str, preview_rows: int = 6, preview_cols: int = 12) -> list[dict]:
    """Return metadata + a tiny preview for every sheet in the workbook.

    Uses read-only mode for speed on huge workbooks. For each sheet we scan the
    first ~30 rows to build a preview and guess where the header row is.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[dict] = []
    try:
        for ws in wb.worksheets:
            scanned: list[list] = []
            non_empty = 0
            max_cols = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                trimmed = _trim_trailing(list(row))
                if _row_has_content(trimmed):
                    non_empty += 1
                    max_cols = max(max_cols, len(trimmed))
                    if i < 30:
                        scanned.append((i, trimmed))
                if i > 400:  # don't walk enormous historical sheets fully
                    break

            header_idx = _guess_header(scanned)
            preview = []
            for idx, trimmed in scanned[:preview_rows]:
                preview.append({
                    "row": idx,
                    "cells": [cell_to_json(c) for c in trimmed[:preview_cols]],
                })

            out.append({
                "name": ws.title,
                "rows": ws.max_row or 0,
                "cols": ws.max_column or 0,
                "non_empty_rows": non_empty,
                "header_guess": header_idx,
                "likely_asset": _looks_like_asset(scanned),
                "preview": preview,
            })
    finally:
        wb.close()
    return out


def _guess_header(scanned: list) -> int | None:
    """Header = first row with >=2 text cells that is followed by a numeric row."""
    for pos, (idx, row) in enumerate(scanned):
        text_cells = sum(1 for c in row if isinstance(c, str) and c.strip())
        if text_cells >= 2:
            for _, nxt in scanned[pos + 1:pos + 4]:
                if any(_is_num(c) for c in nxt):
                    return idx
    return scanned[0][0] if scanned else None


def _looks_like_asset(scanned: list) -> bool:
    """Heuristic: any row mixes a label with a 'largish' number (a value)."""
    for _, row in scanned:
        has_text = any(isinstance(c, str) and c.strip() for c in row)
        has_big_num = any(_is_num(c) and abs(c) >= 1000 for c in row)
        if has_text and has_big_num:
            return True
    return False


# --------------------------------------------------------------------------
# Single-sheet grid (for the mapping UI / import)
# --------------------------------------------------------------------------
def get_grid(path: str, sheet: str, max_rows: int = 300, max_cols: int = 60,
             col_offset: int = 0) -> dict:
    """Return a JSON-safe 2D grid for one sheet, capped to keep payloads sane."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise KeyError(f"sheet {sheet!r} not found")
        ws = wb[sheet]
        total_rows = ws.max_row or 0
        total_cols = ws.max_column or 0

        rows: list[list] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            cells = [cell_to_json(c) for c in row[col_offset:col_offset + max_cols]]
            rows.append(cells)

        # column headers (A, B, C ...) for the slice
        from openpyxl.utils import get_column_letter
        col_letters = [get_column_letter(col_offset + j + 1) for j in range(max_cols)]

        return {
            "sheet": sheet,
            "rows": rows,
            "col_letters": col_letters[:max_cols],
            "total_rows": total_rows,
            "total_cols": total_cols,
            "col_offset": col_offset,
            "truncated_rows": total_rows > max_rows,
            "truncated_cols": total_cols > (col_offset + max_cols),
        }
    finally:
        wb.close()
