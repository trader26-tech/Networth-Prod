"""
Holdings importers — parse a broker's exported file into normalized holdings.

Supported (auto-detected by content, broker hint refines):
  • ICICI Direct  — tab-separated .xls text: Stock Symbol, ISIN Code, Qty, Average Cost Price …
  • IBKR          — Activity Statement .csv: the "Open Positions" section (USD stocks)
  • Motilal Oswal — .xlsx "Equity" sheet: Scrip Name, Quantity, Purchase Price, …, ISIN

Each holding → {name, symbol, isin, exchange, quantity, avg_price, currency, import_price}.
`symbol` is the Yahoo-priceable ticker (NSE symbol resolved from ISIN for Indian
names; the raw US ticker for IBKR). `import_price` is the file's market price,
used as a fallback when a live quote isn't available (e.g. unlisted scrips).
Equity-only exports are used, so bonds are naturally excluded.
"""
from __future__ import annotations

import io
import csv
from typing import Optional

from .motilal_client import _nse_isin_map, _nse_name_map, _norm_name


def _f(v) -> float:
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _sym_from_isin(isin: str) -> Optional[str]:
    if not isin:
        return None
    return _nse_isin_map().get(isin.strip())


# ETFs/liquid funds aren't in NSE's EQUITY_L list — map common ones by name.
_NAME_ALIAS = {
    "NIPPONINDIAETFNIFTYBEES": "NIFTYBEES",
    "NIPPONINDIAETFNIFTY50BEES": "NIFTYBEES",
    "ZERODHAMUTUALFUND": "LIQUIDCASE",
    "ZERODHANIFTY1DRATELIQUIDETF": "LIQUIDCASE",
    "NIPPONINDIAETFGOLDBEES": "GOLDBEES",
    "NIPPONINDIAETFSILVER": "SILVERBEES",
    "NIPPONINDIASILVERETF": "SILVERBEES",
}


def _sym_from_name(name: str) -> Optional[str]:
    """Best-effort NSE symbol from a company name (for ISIN-less exports)."""
    nm = _norm_name(name)
    if not nm:
        return None
    if nm in _NAME_ALIAS:
        return _NAME_ALIAS[nm]
    m = _nse_name_map()
    if nm in m:
        return m[nm]
    if len(nm) >= 8:                       # unambiguous prefix only
        cands = {s for k, s in m.items() if k.startswith(nm)}
        if len(cands) == 1:
            return next(iter(cands))
    return None


# ── ICICI Direct (tab-separated .xls text) ──────────────────────────────────────
def parse_icici(raw: bytes) -> list[dict]:
    text = raw.decode("utf-8", "ignore")
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        raise ValueError("Empty ICICI file.")
    header = [h.strip() for h in lines[0].split("\t")]
    need = {"ISIN Code", "Qty", "Average Cost Price"}
    if not need.issubset(set(header)):
        raise ValueError("Not an ICICI Direct holdings export (missing ISIN/Qty/Average Cost Price columns).")
    idx = {h: i for i, h in enumerate(header)}
    out: list[dict] = []
    for ln in lines[1:]:
        cols = ln.split("\t")
        if len(cols) <= idx["ISIN Code"]:
            continue
        isin = cols[idx["ISIN Code"]].strip()
        qty = _f(cols[idx["Qty"]])
        if not isin or qty <= 0:
            continue
        name = cols[idx.get("Company Name", -1)].strip() if "Company Name" in idx else ""
        ic_sym = cols[idx.get("Stock Symbol", -1)].strip() if "Stock Symbol" in idx else ""
        out.append({
            "name": name, "isin": isin,
            "symbol": _sym_from_isin(isin) or ic_sym,
            "exchange": "NSE", "currency": "INR",
            "quantity": qty, "avg_price": _f(cols[idx["Average Cost Price"]]),
            "import_price": _f(cols[idx["Current Market Price"]]) if "Current Market Price" in idx else 0.0,
        })
    return out


# ── IBKR Activity Statement (.csv) — "Open Positions" section ────────────────────
def parse_ibkr(raw: bytes) -> list[dict]:
    text = raw.decode("utf-8-sig", "ignore")
    rows = list(csv.reader(io.StringIO(text)))
    header = None
    out: list[dict] = []
    for r in rows:
        if not r or r[0] != "Open Positions":
            continue
        if len(r) > 1 and r[1] == "Header":
            header = r
            continue
        if len(r) > 2 and r[1] == "Data" and r[2] == "Summary" and header:
            col = {h: i for i, h in enumerate(header)}
            try:
                if r[col["Asset Category"]].strip().lower() != "stocks":
                    continue
                sym = r[col["Symbol"]].strip()
                qty = _f(r[col["Quantity"]])
                if not sym or qty == 0:
                    continue
                out.append({
                    "name": sym, "isin": "", "symbol": sym, "exchange": "US",
                    "currency": (r[col["Currency"]].strip() or "USD"),
                    "quantity": qty, "avg_price": _f(r[col["Cost Price"]]),
                    "import_price": _f(r[col["Close Price"]]) if "Close Price" in col else 0.0,
                })
            except (KeyError, IndexError):
                continue
    if not out and not any(r and r[0] == "Open Positions" for r in rows):
        raise ValueError("Not an IBKR Activity Statement (no 'Open Positions' section).")
    return out


# ── Motilal Oswal (.xlsx, "Equity" sheet) ───────────────────────────────────────
def parse_motilal(raw: bytes) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = next((s for s in wb.worksheets if s.title.lower().startswith("equity")), wb.worksheets[0])
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = None
    for i, row in enumerate(rows):
        if row and any(str(c).strip() == "Scrip Name" for c in row if c is not None):
            hdr_i = i
            break
    if hdr_i is None:
        raise ValueError("Not a Motilal equity holdings export (no 'Scrip Name' header).")
    header = [str(c).strip() if c is not None else "" for c in rows[hdr_i]]
    idx = {h: i for i, h in enumerate(header)}
    for k in ("Scrip Name", "Quantity", "Purchase Price", "ISIN"):
        if k not in idx:
            raise ValueError(f"Motilal export missing column: {k}")
    out: list[dict] = []
    for row in rows[hdr_i + 1:]:
        if not row or not row[idx["Scrip Name"]]:
            continue
        isin = str(row[idx["ISIN"]] or "").strip()
        qty = _f(row[idx["Quantity"]])
        if qty <= 0:
            continue
        out.append({
            "name": str(row[idx["Scrip Name"]]).strip(), "isin": isin,
            "symbol": _sym_from_isin(isin), "exchange": "NSE", "currency": "INR",
            "quantity": qty, "avg_price": _f(row[idx["Purchase Price"]]),
            "import_price": _f(row[idx["Market Price"]]) if "Market Price" in idx else 0.0,
        })
    return out


# ── Zebu (.xlsx "Summary" sheet — ASSETS table) ─────────────────────────────────
def parse_zebu(raw: bytes) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = next((s for s in wb.worksheets if s.title.lower().startswith("summary")), wb.worksheets[0])
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = None
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in (row or [])]
        if "Scrip" in cells and "N Qty" in cells and "Cl Price" in cells:
            hdr_i = i
            break
    if hdr_i is None:
        raise ValueError("Not a Zebu holdings export (no Scrip / N Qty / Cl Price header).")
    header = [str(c).strip() if c is not None else "" for c in rows[hdr_i]]
    idx = {h: i for i, h in enumerate(header)}
    out: list[dict] = []
    for row in rows[hdr_i + 1:]:
        if not row:
            continue
        scrip = str(row[idx["Scrip"]] or "").strip()
        if not scrip or scrip.lower().startswith("total"):
            break
        qty = _f(row[idx["N Qty"]])
        if qty <= 0:
            continue
        code, _, rest = scrip.partition("-")
        name = rest.strip() if code.strip().isdigit() and rest.strip() else scrip
        out.append({
            "name": name, "isin": "", "symbol": _sym_from_name(name),
            "exchange": "NSE", "currency": "INR",
            "quantity": qty, "avg_price": _f(row[idx["N Rate"]]),
            "import_price": _f(row[idx["Cl Price"]]),
        })
    return out


_PARSERS = {"icici": parse_icici, "ibkr": parse_ibkr, "motilal": parse_motilal, "zebu": parse_zebu}


def parse(broker: str, raw: bytes) -> list[dict]:
    p = _PARSERS.get((broker or "").lower())
    if not p:
        raise ValueError(f"No importer for broker '{broker}'.")
    rows = p(raw)
    if not rows:
        raise ValueError("No holdings found in the file (is it the right export?).")
    return rows
