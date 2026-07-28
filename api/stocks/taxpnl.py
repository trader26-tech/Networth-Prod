"""Parse a Zerodha **Tax P&L** statement (Console → Reports → Tax P&L, the
multi-sheet XLSX) into the realized-gain buckets that matter for filing:

  short_term / long_term  — equity capital gains (STCG @15%, LTCG tax-free ≤₹1.25L)
  intraday                — speculative business income
  fno_options / fno_futures, non_equity — non-speculative business income
  dividends               — taxed at slab (income from other sources)
  charges                 — brokerage + statutory charges

Each file is one CLIENT (Zerodha account) for one FINANCIAL YEAR. The client id
and name come from the sheets; the FY comes from the filename
(`taxpnl-VWM579-2025_2026-Q1-Q4.xlsx` → 2025-2026). Robust to row shifts: every
figure is found by its LABEL, not a fixed cell.
"""
from __future__ import annotations

import io
import re
from typing import Optional


def _find_value(ws, label: str, maxrow: int = 80, maxcol: int = 6):
    """Value in the cell just RIGHT of the first cell whose text starts with
    `label` (case-insensitive). None when the label isn't present."""
    lab = label.strip().lower()
    for r in range(1, min(ws.max_row, maxrow) + 1):
        for c in range(1, min(ws.max_column, maxcol) + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip().lower().startswith(lab):
                return ws.cell(r, c + 1).value
    return None


def _num(ws, label: str) -> float:
    v = _find_value(ws, label)
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return 0.0


def _fy_from_name(filename: str) -> str:
    """`…-2025_2026-…` or `…2025-26…` → 'YYYY-YYYY'. Falls back to ''."""
    s = filename or ""
    m = re.search(r"(20\d{2})[_\-](20\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.search(r"(20\d{2})[_\-](\d{2})\b", s)     # 2025-26
    if m:
        return f"{m.group(1)}-20{m.group(2)}"
    return ""


def parse_tax_pnl(filename: str, raw: bytes) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    # ── client id / name / PAN (from whichever sheet carries them) ──
    client_id = client_name = pan = ""
    for sh in wb.sheetnames:
        ws = wb[sh]
        client_id = client_id or str(_find_value(ws, "Client ID") or "").strip()
        client_name = client_name or str(_find_value(ws, "Client Name") or "").strip()
        pan = pan or str(_find_value(ws, "PAN") or "").strip()
        if client_id and client_name:
            break

    # ── equity + non-equity realized gains ──
    stcg = ltcg = intraday = nonequity = 0.0
    if "Equity and Non Equity" in wb.sheetnames:
        ws = wb["Equity and Non Equity"]
        intraday = _num(ws, "Intraday")
        stcg = _num(ws, "Short Term profit")
        ltcg = _num(ws, "Long Term profit")
        nonequity = _num(ws, "Non Equity profit")

    # ── F&O realized ──
    fno_opt = fno_fut = 0.0
    if "F&O" in wb.sheetnames:
        ws = wb["F&O"]
        fno_opt = _num(ws, "Options Realized")
        fno_fut = _num(ws, "Futures Realized")

    # ── dividends (sum the 'Net Dividend Amount' column under its header) ──
    dividends = 0.0
    if "Equity Dividends" in wb.sheetnames:
        ws = wb["Equity Dividends"]
        amt_col = None
        for r in range(1, min(ws.max_row, 40) + 1):
            for c in range(1, min(ws.max_column, 8) + 1):
                v = ws.cell(r, c).value
                if v and "net dividend" in str(v).strip().lower():
                    amt_col = c
                    hdr_row = r
                    break
            if amt_col:
                break
        if amt_col:
            for r in range(hdr_row + 1, ws.max_row + 1):
                try:
                    dividends += float(ws.cell(r, amt_col).value or 0)
                except (TypeError, ValueError):
                    pass
    dividends = round(dividends, 2)

    return {
        "client_id": client_id, "client_name": client_name, "pan": pan,
        "fy": _fy_from_name(filename),
        "short_term": stcg, "long_term": ltcg, "intraday": intraday,
        "non_equity": nonequity, "fno_options": fno_opt, "fno_futures": fno_fut,
        "dividends": dividends,
        # totals used by the UI
        "equity_gain": round(stcg + ltcg + intraday, 2),
        "fno_gain": round(fno_opt + fno_fut, 2),
        "total_booked": round(stcg + ltcg + intraday + nonequity + fno_opt + fno_fut, 2),
        "filename": filename,
    }


# ── person resolution + storage (KV blob; one entry per client × FY) ──────────
LTCG_EXEMPT = 125000.0        # ₹1.25L/yr equity LTCG is tax-free (current rule)
_TAX_KEY = "tax_pnl_statements"


def resolve_person(client_id: str, client_name: str) -> str:
    """Map a Zerodha client id → the holder's short name. Prefers the person set
    on a connected account with that kite_user_id; else the first name off the
    statement; else the client id."""
    for mod in ("..fno", "..stocks"):
        try:
            if mod == "..fno":
                from ..fno import store as st
                accts = st.list_accounts()
            else:
                from . import store as st
                accts = st.list_accounts() if hasattr(st, "list_accounts") else []
            for a in accts:
                if (a.get("kite_user_id") or a.get("account") or "") == client_id and a.get("person"):
                    return a["person"]
        except Exception:
            pass
    return (client_name or "").split()[0] if client_name else (client_id or "—")


def _kv():
    from ..portfolio import store as kv
    return kv


def save_statement(rec: dict) -> dict:
    """Store (replace) one client×FY statement. Returns the enriched record."""
    rec = dict(rec)
    rec["person"] = resolve_person(rec.get("client_id", ""), rec.get("client_name", ""))
    rec["ltcg_free_left"] = round(max(0.0, LTCG_EXEMPT - float(rec.get("long_term") or 0)), 2)
    key = f"{rec.get('client_id','?')}|{rec.get('fy','?')}"
    try:
        kv = _kv()
        cur = kv.cache_get(_TAX_KEY)
        blob = cur.get("value") if cur and isinstance(cur.get("value"), dict) else {}
        blob[key] = rec
        kv.cache_set(_TAX_KEY, blob)
    except Exception:
        pass
    return rec


def list_statements() -> list:
    try:
        kv = _kv()
        cur = kv.cache_get(_TAX_KEY)
        blob = cur.get("value") if cur and isinstance(cur.get("value"), dict) else {}
        out = []
        for rec in blob.values():
            rec = dict(rec)
            # refresh derived fields (person mapping may have changed)
            rec["person"] = resolve_person(rec.get("client_id", ""), rec.get("client_name", ""))
            rec["ltcg_free_left"] = round(max(0.0, LTCG_EXEMPT - float(rec.get("long_term") or 0)), 2)
            out.append(rec)
        return sorted(out, key=lambda r: (r.get("fy") or "", r.get("person") or ""))
    except Exception:
        return []


def delete_statement(client_id: str, fy: str) -> bool:
    try:
        kv = _kv()
        cur = kv.cache_get(_TAX_KEY)
        blob = cur.get("value") if cur and isinstance(cur.get("value"), dict) else {}
        key = f"{client_id}|{fy}"
        if key in blob:
            blob.pop(key)
            kv.cache_set(_TAX_KEY, blob)
            return True
    except Exception:
        pass
    return False

