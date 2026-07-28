"""
Parse a Zerodha equity tradebook (.xlsx) into individual trades.

Zerodha's layout (per sheet): a few preamble rows (Client ID, period), then a
header row containing 'Symbol', 'Trade Date', 'Trade Type', 'Quantity', 'Price',
'Trade ID', 'Order Execution Time', followed by one row per executed trade.
Every sheet of the workbook is scanned (a workbook can hold multiple segments).
"""
from __future__ import annotations

import re
from typing import Any, Optional

import openpyxl

_WANT = ("symbol", "trade date", "trade type", "quantity", "price")


def _norm(v: Any) -> str:
    return str(v).strip().lower() if v is not None else ""


def _to_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _to_date(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    s = str(v)
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return m.group(0)
    m = re.search(r"(\d{2})[-/](\d{2})[-/](\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return s[:10] or None


def parse_workbook(path: str) -> dict:
    """Return {client_id, period_from, period_to, trades:[...]} from a tradebook."""
    # NOT read_only: some Zerodha exports report a wrong sheet dimension in
    # read-only mode (openpyxl then yields an empty grid).
    wb = openpyxl.load_workbook(path, data_only=True)
    client_id = ""
    period_from = period_to = ""
    trades: list[dict] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        col: dict[str, int] = {}
        for i, r in enumerate(rows):
            cells = [_norm(c) for c in r]
            joined = " ".join(cells)
            if not client_id and "client id" in joined:
                # value sits in the next non-empty cell
                for j, c in enumerate(cells):
                    if c == "client id" and j + 1 < len(r) and r[j + 1]:
                        client_id = str(r[j + 1]).strip()
            if not period_from:
                m = re.search(r"from (\d{4}-\d{2}-\d{2}) to (\d{4}-\d{2}-\d{2})", joined)
                if m:
                    period_from, period_to = m.group(1), m.group(2)
            if all(w in cells for w in _WANT):
                header_idx = i
                for j, c in enumerate(cells):
                    col[c] = j
                break
        if header_idx is None:
            continue

        def cell(r, name):
            j = col.get(name)
            return r[j] if (j is not None and j < len(r)) else None

        for r in rows[header_idx + 1:]:
            sym = cell(r, "symbol")
            if not sym:
                continue
            qty = _to_float(cell(r, "quantity"))
            price = _to_float(cell(r, "price"))
            ttype = _norm(cell(r, "trade type"))
            if qty is None or price is None or ttype not in ("buy", "sell"):
                continue
            trades.append({
                "symbol": str(sym).strip(),
                "isin": (str(cell(r, "isin")).strip() if cell(r, "isin") else None),
                "trade_date": _to_date(cell(r, "trade date")),
                "exchange": (str(cell(r, "exchange")).strip() if cell(r, "exchange") else None),
                "trade_type": ttype,
                "quantity": qty,
                "price": price,
                "trade_id": (str(cell(r, "trade id")).strip() if cell(r, "trade id") else None),
                "order_time": (str(cell(r, "order execution time")).strip()
                               if cell(r, "order execution time") else None),
            })

    return {
        "client_id": client_id,
        "period_from": period_from,
        "period_to": period_to,
        "trades": trades,
    }
