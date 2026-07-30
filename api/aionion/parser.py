"""
Aionion broker holdings-statement (.xlsx) parser.
=================================================

Aionion exports a clean 4-sheet workbook — Portfolio Dashboard, Equity, MF,
Bonds — with real column headers, so unlike a CAS PDF this is read by HEADER
NAME, not fixed positions. That makes it robust to column reordering or extra
columns in future exports.

parse_workbook(path_or_bytes) -> {
    "investor":  {name, pan, client_id, downloaded_on},
    "equities":  [{symbol, isin, quantity, avg_price, market_price, invested,
                   market_value, kind}],   # kind: equity | etf
    "mutual_funds": [{scheme_code, folio, isin, scheme, units, avg_cost,
                      invested, nav, nav_date, value}],
    "bonds":     [{issuer, isin, quantity, invested, value, coupon_rate,
                   maturity_date, call_date, ytm, ytc}],
    "totals":    {equity_market, mf_value, bond_value, net_worth},
    "warnings":  [...],
}

Nothing here is persisted or priced — it returns plain dicts; callers decide
what to store and re-price live.
"""
from __future__ import annotations

import io
import re
from typing import Any


class AionionParseError(Exception):
    """The workbook doesn't look like an Aionion holdings export."""


def _norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def _key(s: Any) -> str:
    """Header → comparable key: upper, collapse non-alnum."""
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def _num(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("Rs.", "").replace("₹", "")
    if s in ("", "-", "'-", "--", "N/A", "NA"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _iso_date(v: Any) -> str | None:
    """Accept '2028-10-27', '29-Jul-2026', datetime → ISO yyyy-mm-dd."""
    if v is None:
        return None
    # openpyxl may hand back a datetime
    if hasattr(v, "isoformat") and hasattr(v, "year"):
        try:
            return v.date().isoformat() if hasattr(v, "date") else v.isoformat()[:10]
        except Exception:
            pass
    s = str(v).strip().strip("'")
    if not s or s in ("-", "--"):
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    m = re.fullmatch(r"(\d{1,2})-([A-Za-z]{3})-(\d{4})", s)
    if m:
        mon = {
            "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
            "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
        }.get(m.group(2).lower())
        if mon:
            return f"{m.group(3)}-{mon:02d}-{int(m.group(1)):02d}"
    return None


def _load(path_or_bytes) -> Any:
    import warnings

    import openpyxl

    src = io.BytesIO(path_or_bytes) if isinstance(path_or_bytes, (bytes, bytearray)) else path_or_bytes
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")  # Aionion adds unsupported xlsx extensions
        return openpyxl.load_workbook(src, read_only=True, data_only=True)


def _rows(ws) -> list[list]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _find_header(rows: list[list], must_have: list[str]) -> int | None:
    """Return the index of the header row containing all `must_have` keys."""
    want = [_key(h) for h in must_have]
    for i, row in enumerate(rows):
        keys = [_key(c) for c in row]
        if all(any(w in k for k in keys) for w in want):
            return i
    return None


def _col_map(header: list) -> dict[str, int]:
    return {_key(c): i for i, c in enumerate(header) if c not in (None, "")}


def _get(row: list, cmap: dict, *aliases: str):
    for a in aliases:
        ka = _key(a)
        for k, i in cmap.items():
            if ka in k and i < len(row):
                return row[i]
    return None


def _classify_equity(symbol: str, isin: str) -> str:
    s = (symbol or "").upper()
    ii = (isin or "").upper()
    if ii.startswith("INF"):
        return "etf"  # ETFs carry INF ISINs like MFs but trade like stocks
    if re.search(r"BEES|ETF|GOLD|LIQUID|NIFTY", s):
        return "etf"
    return "equity"


# ---------------------------------------------------------------------------
def _parse_investor(rows: list[list]) -> dict:
    inv = {"name": None, "pan": None, "client_id": None, "downloaded_on": None}
    for row in rows[:6]:
        cells = [_norm(c) for c in row if c not in (None, "")]
        joined = " | ".join(cells)
        for lbl, key in (("CLIENT ID", "client_id"), ("NAME", "name"), ("PAN", "pan")):
            if cells and _key(cells[0]) == _key(lbl) and len(cells) > 1:
                inv[key] = cells[1]
        m = re.search(r"DOWNLOADED ON:\s*(.+)", joined, re.I)
        if m:
            inv["downloaded_on"] = m.group(1).strip()
    return inv


def _parse_equity(ws, warnings: list) -> list[dict]:
    rows = _rows(ws)
    h = _find_header(rows, ["ISIN", "QTY", "AVG PRICE", "MARKET VALUE"])
    if h is None:
        warnings.append("Equity sheet: header row not found")
        return []
    cmap = _col_map(rows[h])
    out = []
    for row in rows[h + 1:]:
        sym = _norm(_get(row, cmap, "SECURITY", "SYMBOL"))
        isin = _norm(_get(row, cmap, "ISIN", "CODE"))
        if not sym or _key(sym) in ("TOTALS", "TOTAL") or not isin:
            continue
        qty = _num(_get(row, cmap, "QTY", "UNITS"))
        if qty is None:
            continue
        out.append({
            "symbol": sym.split()[0],       # "DRREDDY" (drop any trailing words)
            "isin": isin.upper(),
            "quantity": qty,
            "avg_price": _num(_get(row, cmap, "AVG PRICE")),
            "market_price": _num(_get(row, cmap, "MARKET PRICE")),
            "invested": _num(_get(row, cmap, "INVESTED")),
            "market_value": _num(_get(row, cmap, "MARKET VALUE")),
            "kind": _classify_equity(sym, isin),
        })
    return out


def _parse_mf(ws, warnings: list) -> list[dict]:
    rows = _rows(ws)
    h = _find_header(rows, ["SCHEME NAME", "UNITS", "CURRENT NAV"])
    if h is None:
        warnings.append("MF sheet: header row not found")
        return []
    cmap = _col_map(rows[h])
    out = []
    for row in rows[h + 1:]:
        scheme = _norm(_get(row, cmap, "SCHEME NAME"))
        if not scheme or _key(scheme) in ("TOTALS", "TOTAL"):
            continue
        units = _num(_get(row, cmap, "UNITS"))
        if units is None:
            continue
        # Aionion's MF sheet has no ISIN column; the "FOLIO NO" cell actually
        # carries an ISIN-format value (e.g. INF789F01JN2), and SCHEME CODE is
        # the AMFI code (used for live NAV). Capture both.
        folio_raw = _norm(_get(row, cmap, "FOLIO"))
        isin = folio_raw.upper() if re.fullmatch(r"[A-Z]{2}[A-Z0-9]{9}\d", folio_raw.upper()) else None
        out.append({
            "scheme_code": _norm(_get(row, cmap, "SCHEME CODE")),
            "folio": folio_raw if not isin else None,
            "isin": isin,
            "scheme": scheme,
            "units": units,
            "avg_cost": _num(_get(row, cmap, "AVG COST")),
            "invested": _num(_get(row, cmap, "INVESTED")),
            "nav": _num(_get(row, cmap, "CURRENT NAV")),
            "nav_date": _iso_date(_get(row, cmap, "NAV DATE")),
            "value": _num(_get(row, cmap, "CURRENT VALUE")),
        })
    return out


def _parse_bonds(ws, warnings: list) -> list[dict]:
    rows = _rows(ws)
    h = _find_header(rows, ["BOND NAME", "ISIN", "COUPON", "MATURITY"])
    if h is None:
        warnings.append("Bonds sheet: header row not found")
        return []
    cmap = _col_map(rows[h])
    out = []
    for row in rows[h + 1:]:
        name = _norm(_get(row, cmap, "BOND NAME"))
        isin = _norm(_get(row, cmap, "ISIN"))
        if not name or _key(name) in ("TOTALS", "TOTAL") or not isin:
            continue
        qty = _num(_get(row, cmap, "QTY"))
        # Aionion stores YTM/YTC as a fraction (0.106) — normalise to percent.
        ytm = _num(_get(row, cmap, "YTM"))
        if ytm is not None and ytm < 1:
            ytm *= 100
        ytc = _num(_get(row, cmap, "YTC"))
        if ytc is not None and ytc < 1:
            ytc *= 100
        out.append({
            "issuer": name,
            "isin": isin.upper(),
            "quantity": qty,
            "invested": _num(_get(row, cmap, "INVESTED", "PRINCIPAL")),
            "value": _num(_get(row, cmap, "CURRENT VALUE")),
            "coupon_rate": _num(_get(row, cmap, "COUPON")),
            "maturity_date": _iso_date(_get(row, cmap, "MATURITY")),
            "call_date": _iso_date(_get(row, cmap, "CALL DATE")),
            "ytm": ytm,
            "ytc": ytc,
        })
    return out


def parse_workbook(path_or_bytes) -> dict[str, Any]:
    try:
        wb = _load(path_or_bytes)
    except Exception as e:
        raise AionionParseError(f"Could not open the workbook: {e}") from e

    sheets = {s.lower(): s for s in wb.sheetnames}
    warnings: list[str] = []

    def sheet(*names):
        for n in names:
            if n.lower() in sheets:
                return wb[sheets[n.lower()]]
        return None

    # Investor from the dashboard (or equity) sheet header block
    dash = sheet("Portfolio Dashboard", "Dashboard") or sheet("Equity")
    investor = _parse_investor(_rows(dash)) if dash else {}

    eq_ws, mf_ws, bd_ws = sheet("Equity"), sheet("MF"), sheet("Bonds")
    if not (eq_ws or mf_ws or bd_ws):
        raise AionionParseError(
            "No Equity / MF / Bonds sheet found — is this an Aionion holdings export?"
        )

    equities = _parse_equity(eq_ws, warnings) if eq_ws else []
    mutual_funds = _parse_mf(mf_ws, warnings) if mf_ws else []
    bonds = _parse_bonds(bd_ws, warnings) if bd_ws else []

    totals = {
        "equity_market": round(sum(e.get("market_value") or 0 for e in equities), 2),
        "mf_value": round(sum(m.get("value") or 0 for m in mutual_funds), 2),
        "bond_value": round(sum(b.get("value") or 0 for b in bonds), 2),
    }
    totals["net_worth"] = round(
        totals["equity_market"] + totals["mf_value"] + totals["bond_value"], 2
    )

    return {
        "investor": investor,
        "equities": equities,
        "mutual_funds": mutual_funds,
        "bonds": bonds,
        "totals": totals,
        "warnings": warnings,
    }
