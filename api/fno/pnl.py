"""
F&O P&L math that doesn't need Kite live data.

  rebuild_daily_from_trades — replay the fill log (avg-cost netting per
      instrument) and write realised P&L per account × day × strategy. Used to
      backfill history (e.g. the Console tradebook import); it never overwrites
      a 'live' row written from Kite positions during market hours.

  parse_tradebook — Zerodha Console → Reports → Tradebook export (CSV or XLSX,
      F&O / MCX segment) → normalized fno_trades rows.
"""
from __future__ import annotations

import csv
import io
from collections import defaultdict
from typing import Optional

from . import store, kite as fno_kite


def rebuild_daily_from_trades(account_id: Optional[str] = None) -> dict:
    """Replay fills chronologically; realised P&L lands on the closing fill's
    date (matches how Console's P&L report attributes closed trades)."""
    # Days an imported P&L statement covers are AUTHORITATIVE — the fill replay
    # (which can be incomplete) neither clobbers nor double-counts them: it skips
    # those days entirely and only fills the gaps (recent days no statement covers
    # yet). delete_daily_before already preserves source='statement' rows.
    windows = store.statement_windows(account_id) if account_id else []
    trades = store.all_trades()
    if account_id:
        trades = [t for t in trades if t.get("account_id") == account_id]
        # A rebuild is authoritative for COMPLETED days → clear any stale daily
        # rows for past days first (incl. old 'live' rows that only ever held
        # unrealized m2m with realized=0, which would otherwise blank the day).
        # Today's live row is left alone (date == today is not < today).
        store.delete_daily_before(account_id, store.today_ist())
    trades.sort(key=lambda t: (t.get("fill_ts") or t.get("trade_date") or ""))
    overrides = store.get_account_strategies()   # per-account strategy pins
    trade_pins = store.get_trade_strategies()    # per-trade overrides (win over the pin)

    # per (account, instrument): open position {qty (signed), avg}
    pos: dict = defaultdict(lambda: {"qty": 0.0, "avg": 0.0})
    # per (account, date, strategy): {realized, count}
    days: dict = defaultdict(lambda: {"realized": 0.0, "count": 0})

    for t in trades:
        acc = t.get("account_id")
        sym = t.get("tradingsymbol") or ""
        key = (acc, sym)
        strategy = store.resolve_trade_strategy(t, overrides, trade_pins)
        date = (t.get("trade_date") or "")[:10]
        mult = fno_kite.contract_multiplier(t.get("exchange"), sym)  # ₹/point (crude=100)
        qty = float(t.get("quantity") or 0)
        price = float(t.get("price") or 0)
        side = 1.0 if (t.get("transaction_type") or "").upper() == "BUY" else -1.0
        signed = qty * side

        p = pos[key]
        days[(acc, date, strategy)]["count"] += 1

        if p["qty"] == 0 or (p["qty"] > 0) == (signed > 0):
            # opening / extending — new weighted average
            total = abs(p["qty"]) + qty
            p["avg"] = (abs(p["qty"]) * p["avg"] + qty * price) / total if total else 0.0
            p["qty"] += signed
            continue

        # reducing (or flipping) an open position → realise on the closed part
        closed = min(abs(p["qty"]), qty)
        direction = 1.0 if p["qty"] > 0 else -1.0   # long: sell above avg wins
        days[(acc, date, strategy)]["realized"] += (price - p["avg"]) * closed * direction * mult
        p["qty"] += signed
        if (p["qty"] > 0) == (signed > 0) and p["qty"] != 0:
            p["avg"] = price                          # flipped — remainder opens here
        elif p["qty"] == 0:
            p["avg"] = 0.0

    def _covered(d: str) -> bool:                  # day owned by a P&L statement
        return any(f <= d <= t for f, t in windows)

    for (acc, date, strategy), v in days.items():
        if not date or _covered(date):
            continue                                # statement is authoritative here
        store.upsert_daily(acc, date, strategy, realized=round(v["realized"], 2),
                           unrealized=0.0, total=round(v["realized"], 2),
                           trades_count=v["count"], source="trades")

    open_syms = [f"{k[1]} ({int(p['qty'])})" for k, p in pos.items() if abs(p["qty"]) > 1e-9]
    return {"days_written": len(days), "open_positions": open_syms}


def realized_by_strategy(account_id: str, day: str) -> dict:
    """Fill-replay realised P&L for ONE account on ONE day, per strategy — the
    AUTHORITATIVE 'booked today'. The live engine can't be trusted for this:
    Kite's position `realised` field reports 0 for a position carried from a
    previous session and closed today, so the daily row's realised comes out 0
    even though the fills clearly booked a profit/loss. Replaying the fill log
    (same avg-cost netting + lot multiplier as rebuild_daily_from_trades) gets
    it right and keeps 'Booked today' consistent with the trade history."""
    trades = [t for t in store.all_trades() if t.get("account_id") == account_id]
    trades.sort(key=lambda t: (t.get("fill_ts") or t.get("trade_date") or ""))
    overrides = store.get_account_strategies()
    trade_pins = store.get_trade_strategies()
    pos: dict = defaultdict(lambda: {"qty": 0.0, "avg": 0.0})
    out: dict = defaultdict(float)
    for t in trades:
        sym = t.get("tradingsymbol") or ""
        strategy = store.resolve_trade_strategy(t, overrides, trade_pins)
        d = (t.get("trade_date") or "")[:10]
        mult = fno_kite.contract_multiplier(t.get("exchange"), sym)
        qty = float(t.get("quantity") or 0)
        price = float(t.get("price") or 0)
        side = 1.0 if (t.get("transaction_type") or "").upper() == "BUY" else -1.0
        signed = qty * side
        p = pos[sym]
        if p["qty"] == 0 or (p["qty"] > 0) == (signed > 0):
            total = abs(p["qty"]) + qty
            p["avg"] = (abs(p["qty"]) * p["avg"] + qty * price) / total if total else 0.0
            p["qty"] += signed
            continue
        closed = min(abs(p["qty"]), qty)
        direction = 1.0 if p["qty"] > 0 else -1.0
        if d == day:
            out[strategy] += (price - p["avg"]) * closed * direction * mult
        p["qty"] += signed
        if (p["qty"] > 0) == (signed > 0) and p["qty"] != 0:
            p["avg"] = price
        elif p["qty"] == 0:
            p["avg"] = 0.0
    return {k: round(v, 2) for k, v in out.items()}


# ── Zerodha P&L statement import (Reports → P&L → Download) ──────────────────────
# The tradebook (fills) can be incomplete — Kite caps exports at ~1 year, and a
# missing period corrupts the avg-cost replay. The P&L STATEMENT instead carries
# Zerodha's own AUTHORITATIVE realised + unrealised P&L per symbol (exactly what
# Console shows), so importing it fixes the booked total outright. It has no
# per-fill dates, so realised is attributed to each option's EXPIRY (capped at
# the statement's end date) — the daily totals then still reconcile to the top.
def _num(x):
    try:
        return float(str(x).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _stmt_exchange(sym: str) -> str:
    """Infer the exchange from a bare P&L-statement symbol (no exchange column)."""
    s = (sym or "").upper()
    if s.startswith(("CRUDEOIL", "NATURALGAS", "GOLD", "SILVER", "COPPER", "ZINC",
                     "LEAD", "ALUMINI", "NICKEL", "MENTHAOIL")):
        return "MCX"
    if s.startswith(("SENSEX", "BANKEX")):
        return "BFO"
    return "NFO"


def parse_pnl_statement(account_id: str, filename: str, raw: bytes) -> dict:
    """Zerodha P&L statement (XLSX) → {summary, symbols[], date_from, date_to}.
    Reads every sheet's per-symbol table + the summary block."""
    import openpyxl
    if not (filename or "").lower().endswith((".xlsx", ".xls")):
        raise ValueError("The P&L statement is an Excel file — download it from "
                         "Console → Reports → P&L and upload the .xlsx as-is.")
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    summary = {"realized": 0.0, "unrealized": 0.0, "charges": 0.0, "other": 0.0}
    symbols: list[dict] = []
    date_from = date_to = None
    import re as _re
    for sn in wb.sheetnames:
        grid = [list(r) for r in wb[sn].iter_rows(values_only=True)]
        hdr, col = None, {}
        for i, row in enumerate(grid):
            cells = [str(c or "").strip() for c in row]
            low = [c.lower() for c in cells]
            for j, c in enumerate(low):                         # summary label/value pairs
                if hdr is None and c in ("realized p&l", "unrealized p&l", "charges",
                                         "other credit & debit") and j + 1 < len(row):
                    v = _num(row[j + 1])
                    if v is not None:
                        key = {"realized p&l": "realized", "unrealized p&l": "unrealized",
                               "charges": "charges", "other credit & debit": "other"}[c]
                        summary[key] += v
            if date_from is None:
                m = _re.search(r"from\s+(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})",
                               " ".join(cells), _re.I)
                if m:
                    date_from, date_to = m.group(1), m.group(2)
            if hdr is None and "symbol" in low and "realized p&l" in low and "quantity" in low:
                hdr = i
                col = {name: low.index(name) for name in low if name}
        if hdr is None:
            continue

        def g(row, name):
            idx = col.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        for row in grid[hdr + 1:]:
            sym = str(g(row, "symbol") or "").strip().upper()
            if not sym or sym == "TOTAL":
                continue
            symbols.append({
                "symbol": sym, "exchange": _stmt_exchange(sym),
                "qty": _num(g(row, "quantity")) or 0.0,
                "buy_value": _num(g(row, "buy value")) or 0.0,
                "sell_value": _num(g(row, "sell value")) or 0.0,
                "realized": _num(g(row, "realized p&l")) or 0.0,
                "open_qty": _num(g(row, "open quantity")) or 0.0,
                "open_side": str(g(row, "open quantity type") or "").strip().lower(),
                "unrealized": _num(g(row, "unrealized p&l")) or 0.0,
            })
    if not symbols:
        raise ValueError("Couldn't find the per-symbol P&L table — download the "
                         "F&O P&L statement from Console → Reports → P&L (Excel).")
    return {"summary": {k: round(v, 2) for k, v in summary.items()},
            "symbols": symbols, "date_from": date_from, "date_to": date_to}


def statement_days(account_id: str, parsed: dict) -> list[dict]:
    """Per-(date, strategy) realised a single statement contributes: each symbol's
    realised booked on its expiry, clamped into the statement window and capped at
    yesterday (today belongs to the live engine). Pure — no DB writes. Stored on
    the statement rec so the authoritative daily rows can be re-derived from ALL
    statements after any add/remove."""
    from . import openpos as fno_openpos
    from datetime import date as _date, timedelta as _td
    _y = (_date.fromisoformat(store.today_ist()) - _td(days=1)).isoformat()
    end = min(parsed.get("date_to") or _y, _y)
    start = min(parsed.get("date_from") or end, end)
    overrides = store.get_account_strategies()
    days: dict = defaultdict(lambda: {"realized": 0.0, "count": 0})
    for s in parsed["symbols"]:
        if abs(s["realized"]) < 1e-9:
            continue
        sym, ex = s["symbol"], s["exchange"]
        strat = store.resolve_trade_strategy(
            {"tradingsymbol": sym, "exchange": ex, "strategy": fno_kite.classify(ex, sym)},
            overrides, {})
        exp = fno_openpos.option_expiry(sym)
        d = exp.isoformat() if exp else end
        d = min(max(d, start), end)                    # clamp into the statement window
        days[(d, strat)]["realized"] += s["realized"]
        days[(d, strat)]["count"] += 1
    return [{"date": d, "strategy": strat, "realized": round(v["realized"], 2), "count": v["count"]}
            for (d, strat), v in days.items()]


def rebuild_statement_daily(account_id: str) -> dict:
    """Rewrite the AUTHORITATIVE source='statement' daily rows from the UNION of
    every imported statement (windows are kept disjoint on import → summing is
    safe). Then replay fills for the days NO statement covers, so booked stays
    complete without ever double-counting a statement period."""
    store.delete_daily_by_source(account_id, "statement")
    agg: dict = defaultdict(lambda: {"realized": 0.0, "count": 0})
    for st in store.get_pnl_statements(account_id):
        for row in st.get("days", []):
            d, strat = row.get("date"), row.get("strategy")
            if not d:
                continue
            agg[(d, strat)]["realized"] += float(row.get("realized") or 0)
            agg[(d, strat)]["count"] += int(row.get("count") or 0)
    for (d, strat), v in agg.items():
        store.upsert_daily(account_id, d, strat, realized=round(v["realized"], 2),
                           unrealized=0.0, total=round(v["realized"], 2),
                           trades_count=v["count"], source="statement")
    rebuild_daily_from_trades(account_id)              # fill the days no statement covers
    return {"days_written": len(agg), "statements": len(store.get_pnl_statements(account_id))}


# ── Console tradebook import ────────────────────────────────────────────────────
_HDR_ALIASES = {
    "symbol": "tradingsymbol", "tradingsymbol": "tradingsymbol",
    "trade_date": "trade_date", "exchange": "exchange", "segment": "segment",
    "trade_type": "transaction_type", "quantity": "quantity", "qty": "quantity",
    "price": "price", "trade_id": "trade_id", "order_id": "order_id",
    "order_execution_time": "fill_ts", "trade_time": "fill_ts",
    "expiry_date": "expiry", "instrument_type": "instrument_type",
}


def _norm_exchange(exchange: str, segment: str) -> str:
    """Console labels the exchange column with the parent venue (NSE / BSE / MCX)
    and the derivative type in `segment` (FO / CD / COM). Map that pair to the
    Kite exchange code we classify + store on (NFO / BFO / MCX / CDS / BCD)."""
    ex = (exchange or "").strip().upper()
    seg = (segment or "").strip().upper()
    if ex in ("NFO", "BFO", "MCX", "CDS", "BCD"):
        return ex
    fo = seg in ("FO", "FUT", "OPT", "FUTIDX", "OPTIDX", "FUTSTK", "OPTSTK")
    cur = seg in ("CD", "CDS", "CUR", "CURRENCY")
    com = seg in ("COM", "COMM", "COMMODITY", "MCX", "FUTCOM", "OPTFUT")
    if ex == "NSE":
        return "NFO" if fo else "CDS" if cur else ex
    if ex == "BSE":
        return "BFO" if fo else "BCD" if cur else ex
    if ex in ("MCX", "NCDEX"):
        return "MCX"
    # exchange column blank/unknown → fall back to the segment alone
    return "MCX" if com else "NFO" if fo else "CDS" if cur else ex


def _ist(fill) -> Optional[str]:
    """Normalize a fill timestamp to an explicit IST ISO string. Console times
    are IST but tz-naive; without a zone a TIMESTAMPTZ column would read them as
    UTC and shift every trade by 5h30m."""
    if fill is None:
        return None
    s = fill.isoformat() if hasattr(fill, "isoformat") else str(fill).strip()
    if not s:
        return None
    if "T" not in s and " " in s:
        s = s.replace(" ", "T", 1)
    has_tz = s.endswith("Z") or ("+" in s[10:]) or ("-" in s[11:])
    return s if has_tz else s + "+05:30"


def _rows_from_file(filename: str, raw: bytes) -> list[list]:
    if (filename or "").lower().endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [[c for c in row] for row in ws.iter_rows(values_only=True)]
    text = raw.decode("utf-8-sig", errors="replace")
    return [r for r in csv.reader(io.StringIO(text))]


def parse_tradebook(account_id: str, filename: str, raw: bytes, source: str = "import") -> list[dict]:
    """Console tradebook (any segment; non-derivative rows are dropped)."""
    grid = _rows_from_file(filename, raw)
    # Console XLSX files carry a title block — find the real header row.
    hdr_idx, hdr_map = None, {}
    for i, row in enumerate(grid[:40]):
        cells = [str(c or "").strip().lower() for c in row]
        if "trade_id" in cells and ("symbol" in cells or "tradingsymbol" in cells):
            hdr_idx = i
            hdr_map = {j: _HDR_ALIASES[c] for j, c in enumerate(cells) if c in _HDR_ALIASES}
            break
    if hdr_idx is None:
        raise ValueError("Couldn't find the tradebook header row — export the "
                         "Tradebook (F&O segment) from Zerodha Console and upload it as-is.")

    out: list[dict] = []
    for row in grid[hdr_idx + 1:]:
        rec = {name: row[j] if j < len(row) else None for j, name in hdr_map.items()}
        sym = str(rec.get("tradingsymbol") or "").strip().upper()
        trade_id = str(rec.get("trade_id") or "").strip()
        if not sym or not trade_id or trade_id.lower() == "none":
            continue
        ex = _norm_exchange(rec.get("exchange"), rec.get("segment"))
        if ex not in ("NFO", "MCX", "BFO", "CDS", "BCD"):
            continue
        txn = str(rec.get("transaction_type") or "").strip().upper()
        if txn not in ("BUY", "SELL"):
            continue
        date = str(rec.get("trade_date") or "").strip()[:10]
        fill_iso = _ist(rec.get("fill_ts"))
        itype = str(rec.get("instrument_type") or "").strip().upper() or (
            "CE" if sym.endswith("CE") else "PE" if sym.endswith("PE") else
            "FUT" if sym.endswith("FUT") else "")
        try:
            qty = abs(float(str(rec.get("quantity")).replace(",", "")))
            price = float(str(rec.get("price")).replace(",", ""))
        except (TypeError, ValueError):
            continue
        out.append({
            "account_id": account_id,
            "trade_id": trade_id,
            "order_id": str(rec.get("order_id") or ""),
            "strategy": fno_kite.classify(ex, sym, itype),
            "tradingsymbol": sym,
            "exchange": ex,
            "instrument_type": itype,
            "transaction_type": txn,
            "quantity": qty,
            "price": price,
            "product": None,
            "trade_date": date,
            "fill_ts": fill_iso,
            "source": source,
        })
    return out
