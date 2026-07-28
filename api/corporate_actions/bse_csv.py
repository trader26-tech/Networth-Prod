"""
BSE / NSE "Corporate Actions" CSV → declared corporate actions.

The exchange's Corporate Actions export (Security Code, Security Name, Company
Name, Ex Date, Purpose, Record Date, … , Actual Payment Date) lists every
declared action for a window — dividends, bonuses, splits, buybacks, rights.
We parse ALL of them so the upload history can show the full picture, but only
DIVIDEND rows carry a ₹/share amount and feed the dividend log; the rest are
recorded for reference.

Dividend rows come out in the SAME shape as ``nse.fetch_dividends`` so the
existing ingester consumes them unchanged:
    {symbol, ex_date, per_share, subject, record_date, name, code}
plus a ``kind`` on every parsed row (dividend | bonus | split | buyback |
rights | merger | other) and the raw ``purpose``.

This is the manual path for cloud deploys, where NSE's live feed 403s a
datacenter IP — you download the CSV once and upload it here.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from typing import Optional

_NUM = r"(\d+(?:\.\d+)?)"


def _parse_date(s: str) -> Optional[str]:
    """'10 Jul 2026' (and a few common variants) → ISO 'YYYY-MM-DD'."""
    s = (s or "").strip()
    if not s or s == "-":
        return None
    for fmt in ("%d %b %Y", "%d-%b-%Y", "%d %B %Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _classify(purpose: str) -> str:
    p = (purpose or "").lower()
    if "dividend" in p or "income distribution" in p:
        return "dividend"
    if "bonus" in p:
        return "bonus"
    if "split" in p:
        return "split"
    if "buy back" in p or "buyback" in p:
        return "buyback"
    if "right" in p:
        return "rights"
    if "amalgamation" in p or "merger" in p or "scheme of arrangement" in p:
        return "merger"
    return "other"


def _parse_amount(purpose: str) -> Optional[float]:
    """₹/share from a purpose line. BSE writes 'Final Dividend - Rs. - 6.7000';
    a single filing can bundle components ('Special … Rs 2 & Final … Rs 5' or two
    rows), each parsed on its own row. Sums every 'Rs <n>' it finds on the line.
    Returns None for a dividend with no amount stated (e.g. 'Interim Dividend')."""
    s = purpose or ""
    if "dividend" not in s.lower():
        return None
    total, found = 0.0, False
    # 'Rs. - 6.70', 'Rs - 6.70', 'Rs 6.70', 'Re. 1', '₹6.70', 'INR 6.70'
    for m in re.finditer(r"(?:Rs\.?|Re\.?|₹|INR)\s*-?\s*" + _NUM, s, re.I):
        total += float(m.group(1)); found = True
    return round(total, 4) if (found and total > 0) else None


# header aliases → our field name (lower-cased, stripped, trailing punctuation gone)
_HDR = {
    "security code": "code", "security name": "symbol", "company name": "name",
    "ex date": "ex_date", "purpose": "purpose", "record date": "record_date",
    "actual payment date": "payment_date",
}


def _grid(filename: str, raw: bytes) -> list[list]:
    if (filename or "").lower().endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [[c for c in row] for row in ws.iter_rows(values_only=True)]
    text = raw.decode("utf-8-sig", errors="replace")
    return [r for r in csv.reader(io.StringIO(text))]


def parse(filename: str, raw: bytes) -> dict:
    """Parse a Corporate Actions export → {rows, dividends, kinds, date_from,
    date_to}. `dividends` is the ingester-ready subset (kind == dividend AND a
    parseable ₹/share); `rows` is every parsed action for the upload history."""
    grid = _grid(filename, raw)
    if not grid:
        raise ValueError("The file is empty.")

    # find the header row (title blocks sometimes precede it) + column map
    hdr_idx, cols = None, {}
    for i, row in enumerate(grid[:20]):
        cells = [re.sub(r"[\s ]+", " ", str(c or "")).strip().strip(",").lower()
                 for c in row]
        if "purpose" in cells and any(c in ("security name", "company name") for c in cells):
            hdr_idx = i
            cols = {j: _HDR[c] for j, c in enumerate(cells) if c in _HDR}
            break
    if hdr_idx is None:
        raise ValueError("Couldn't find the header row — upload the exchange's "
                         "Corporate Actions export (with a 'Purpose' column) as-is.")

    rows: list[dict] = []
    dividends: list[dict] = []
    kinds: dict[str, int] = {}
    dates: list[str] = []

    for raw_row in grid[hdr_idx + 1:]:
        rec = {name: raw_row[j] if j < len(raw_row) else None for j, name in cols.items()}
        symbol = str(rec.get("symbol") or "").strip().upper()
        purpose = re.sub(r"[\s ]+", " ", str(rec.get("purpose") or "")).strip()
        if not symbol or not purpose:
            continue
        ex = _parse_date(str(rec.get("ex_date") or ""))
        rec_date = _parse_date(str(rec.get("record_date") or ""))
        kind = _classify(purpose)
        per_share = _parse_amount(purpose) if kind == "dividend" else None
        row = {
            "code": str(rec.get("code") or "").strip(),
            "symbol": symbol,
            "name": str(rec.get("name") or "").strip() or symbol,
            "ex_date": ex, "record_date": rec_date,
            "payment_date": _parse_date(str(rec.get("payment_date") or "")),
            "purpose": purpose, "kind": kind, "per_share": per_share,
        }
        rows.append(row)
        kinds[kind] = kinds.get(kind, 0) + 1
        if ex:
            dates.append(ex)
        if kind == "dividend" and per_share and ex:
            dividends.append({
                "symbol": symbol, "ex_date": ex, "per_share": per_share,
                "subject": purpose, "record_date": rec_date,
                "name": row["name"], "code": row["code"],
            })

    if not rows:
        raise ValueError("No corporate actions found in that file.")
    return {
        "rows": rows, "dividends": dividends, "kinds": kinds,
        "date_from": min(dates) if dates else None,
        "date_to": max(dates) if dates else None,
    }
